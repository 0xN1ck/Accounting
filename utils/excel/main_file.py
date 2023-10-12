import os
import shutil
from openpyxl import load_workbook, worksheet
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import coordinate_to_tuple


def create_new_month(self, flag=False):
    """
        Функция create_new_month создает новый файл Excel для текущего месяца. Если файл уже существует,
    то выводится сообщение об ошибке. Функция копирует шаблон файла и переименовывает его в соответствии с выбранным
    пользователем и временными интервалами. Затем сохраняет файл и выводит сообщение об успешном создании файла.

    :param self:
    :param flag:
    :return:
    """
    current_datetime = self.ui.date_start.dateTime().toPyDateTime()
    filename = './docs/' + current_datetime.strftime("%B %Y.xlsx")

    if os.path.exists(filename):
        self.label_status.setText(f'<font color="red">Файл уже создан </font>')
        return
    shutil.copyfile("docs/Шаблон.xlsx", filename)
    workbook = load_workbook(filename)
    template_sheet = workbook["Шаблон"]

    new_sheet = workbook.copy_worksheet(template_sheet)
    new_sheet.title = f'{self.ui.cb_user.currentText()}' \
                      f'({self.ui.date_start.dateTime().toString("d HH-mm")}-' \
                      f'{self.ui.date_finish.dateTime().toString("d HH-mm")})'

    workbook.active = len(workbook.sheetnames) - 1

    template_sheet.sheet_view.tabSelected = False
    template_sheet.sheet_state = "hidden"

    workbook.save(filename)

    if flag:
        flag = False
        return {
            'filename': filename,
            'current_sheet': new_sheet,
            'workbook': workbook,
        }

    self.label_status.setText(f'<font color="green">Файл создан </font>')


def check_current_result_file(self):
    """
            Функция check_current_result_file проверяет существование текущего файла результатов. Если файл уже
        существует, то загружает его и проверяет наличие листа с именем, соответствующим выбранным параметрам.
        Если лист не найден, выводится сообщение об ошибке. Если файл не существует, вызывается функция
        create_new_month для создания нового файла. В результате возвращается словарь с информацией
        о файле и текущем листе.

    :param self:
    :return:
    """
    current_datetime = self.ui.date_start.dateTime().toPyDateTime()
    filename = self.path_all['file']  # + current_datetime.strftime("%B %Y.xlsx")

    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet_name = f'{self.ui.cb_user.currentText()}' \
                     f'({self.ui.date_start.dateTime().toString("d HH-mm")}-' \
                     f'{self.ui.date_finish.dateTime().toString("d HH-mm")})'

        if not self.ui.check_box_list.isChecked():
            result = {
                'filename': filename,
                'current_sheet': (lambda: workbook[sheet_name] if sheet_name in workbook else None)(),
                'workbook': workbook,
            }
            if result['current_sheet'] is None:
                self.label_status.setText(f'<font color="red">Не найден лист {sheet_name}</font>')
                return 0
            self.path_all['file'] = filename
            self.ui.le_file.setText(filename.split('/')[-1])
            return result

        self.ui.check_box_list.setChecked(False)

        if sheet_name in workbook.sheetnames:
            existing_sheet = workbook[sheet_name]
            workbook.remove(existing_sheet)

        template_sheet = workbook["Шаблон"]
        new_sheet = workbook.copy_worksheet(template_sheet)
        new_sheet.title = sheet_name
        workbook.active = len(workbook.sheetnames) - 1

        template_sheet.sheet_view.tabSelected = False
        template_sheet.sheet_state = "hidden"

        for sheet in workbook:
            if sheet.title == sheet_name:
                sheet.sheet_view.tabSelected = True
            else:
                sheet.sheet_view.tabSelected = False

        workbook.save(filename)

    else:
        results = create_new_month(self, flag=True)
        new_sheet = results['current_sheet']
        workbook = results['workbook']

    result = {
        'filename': filename,
        'current_sheet': new_sheet,
        'workbook': workbook,
    }

    self.path_all['file'] = filename
    self.ui.le_file.setText(filename.split('/')[-1])

    return result


def find_empty_cell_top(row_index, column_index, sheet):
    """
            Функция find_empty_cell_top ищет первую пустую ячейку сверху вниз, начиная с указанной строки и столбца
        в заданном листе.

    :param row_index:
    :param column_index:
    :param sheet:
    :return:
    """
    # Поиск первой пустой ячейки сверху вниз
    empty_cell = None
    # row_index = 3

    while empty_cell is None:
        cell_value = sheet.cell(row=row_index, column=column_index).value
        if cell_value is None:
            empty_cell = sheet.cell(row=row_index, column=column_index).coordinate
        else:
            row_index += 1

    return coordinate_to_tuple(empty_cell)


def find_empty_cell_bottom(row_index, column_index, sheet):
    """
            Функция find_empty_cell_bottom ищет первую пустую ячейку снизу вверх, начиная с указанной строки и столбца
        в заданном листе.

    :param row_index:
    :param column_index:
    :param sheet:
    :return:
    """
    # Поиск первой пустой ячейки сверху вниз
    empty_cell = None

    while empty_cell is None:
        cell_value = sheet.cell(row=row_index, column=column_index).value
        if cell_value is None:
            empty_cell = sheet.cell(row=row_index, column=column_index).coordinate
        else:
            row_index -= 1

    return coordinate_to_tuple(empty_cell)


def set_cell(cell, value):
    """
    Функция set_cell устанавливает значение ячейки, задает шрифт и выравнивание для указанной ячейки.

    :param cell:
    :param value:
    :return:
    """
    font = Font(name='Arial', size=12, bold=False, italic=False)
    alignment = Alignment(horizontal='center')
    cell.value = value
    cell.font = font
    cell.alignment = alignment