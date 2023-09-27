import os
import shutil
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import coordinate_to_tuple


def create_new_month(self, flag=False):
    current_datetime = self.ui.date_start.dateTime().toPyDateTime()
    filename = './docs/' + current_datetime.strftime("%B %Y.xlsx")

    shutil.copyfile("docs/Шаблон.xlsx", filename)
    workbook = load_workbook(filename)
    template_sheet = workbook["Шаблон"]

    new_sheet = workbook.copy_worksheet(template_sheet)
    new_sheet.title = f'{self.ui.cb_user.currentText()}' \
                      f'({self.ui.date_start.dateTime().toString("d HH-mm")}-' \
                      f'{self.ui.date_finish.dateTime().toString("d HH-mm")})'
    workbook.active = len(workbook.sheetnames) - 1

    template_sheet.sheet_state = "hidden"

    workbook.save(filename)

    if flag:
        flag = False
        return {
            'filename': filename,
            'current_sheet': new_sheet,
            'workbook': workbook,
        }


def check_current_result_file(self):
    current_datetime = self.ui.date_start.dateTime().toPyDateTime()
    filename = './docs/' + current_datetime.strftime("%B %Y.xlsx")

    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet_name = f'{self.ui.cb_user.currentText()}' \
                     f'({self.ui.date_start.dateTime().toString("d HH-mm")}-' \
                     f'{self.ui.date_finish.dateTime().toString("d HH-mm")})'

        if not self.ui.check_box_list.isChecked():
            result = {
                'filename': filename,
                'current_sheet': (lambda: workbook[sheet_name] if sheet_name in workbook else None)(),
                'workbook': workbook,
            }
            if result['current_sheet'] is None:
                print(f'Не найден лист {sheet_name}')
                return 0
            self.path_all['file'] = filename
            self.ui.le_file.setText(filename.split('/')[-1])
            return result

        self.ui.check_box_list.setChecked(False)

        if sheet_name in workbook.sheetnames:
            existing_sheet = workbook[sheet_name]
            workbook.remove(existing_sheet)

        template_sheet = workbook["Шаблон"]
        new_sheet = workbook.copy_worksheet(template_sheet)
        new_sheet.title = sheet_name
        workbook.active = len(workbook.sheetnames) - 1

        template_sheet.sheet_state = "hidden"

        workbook.save(filename)

    else:
        results = create_new_month(self, flag=True)
        new_sheet = results['current_sheet']
        workbook = results['workbook']

    result = {
        'filename': filename,
        'current_sheet': new_sheet,
        'workbook': workbook,
    }

    self.path_all['file'] = filename
    self.ui.le_file.setText(filename.split('/')[-1])

    return result


def find_empty_cell_top(column_index, sheet):
    # Поиск первой пустой ячейки сверху вниз
    empty_cell = None
    row_index = 3

    while empty_cell is None:
        cell_value = sheet.cell(row=row_index, column=column_index).value
        if cell_value is None:
            empty_cell = sheet.cell(row=row_index, column=column_index).coordinate
        else:
            row_index += 1

    return coordinate_to_tuple(empty_cell)


def find_empty_cell_bottom(column_index, sheet):
    # Поиск первой пустой ячейки сверху вниз
    empty_cell = None
    row_index = 194

    while empty_cell is None:
        cell_value = sheet.cell(row=row_index, column=column_index).value
        if cell_value is None:
            empty_cell = sheet.cell(row=row_index, column=column_index).coordinate
        else:
            row_index -= 1

    return coordinate_to_tuple(empty_cell)


def set_cell(cell, value):
    font = Font(name='Arial', size=12, bold=False, italic=False)
    alignment = Alignment(horizontal='center')
    cell.value = value
    cell.font = font
    cell.alignment = alignment


def write_to_excel_gara_history_p2p(filename, sheet, workbook, filtered_data_gara_history_p2p):
    cell_otc_gar = find_empty_cell_top(2, sheet)
    cell_turnover = find_empty_cell_top(1, sheet)

    for index, (value_res, value_total) in enumerate(
            zip(filtered_data_gara_history_p2p['Результат'], filtered_data_gara_history_p2p['Итого'])):
        cell_res = sheet.cell(row=cell_otc_gar[0] + index, column=cell_otc_gar[1])
        set_cell(cell_res, value_res)

        cell_total = sheet.cell(row=cell_turnover[0] + index, column=cell_turnover[1])
        set_cell(cell_total, value_total)

    workbook.save(filename)


def write_to_excel_gara_swap_history(filename, sheet, workbook, filtered_data_gara_history_swap):
    cell_buy_gar_rub = find_empty_cell_top(5, sheet)
    cell_buy_gar_usdt = find_empty_cell_top(7, sheet)
    cell_buy_gar_course = find_empty_cell_top(6, sheet)
    cell_buy_gar_market = find_empty_cell_top(4, sheet)

    # Запись данных из DataFrame в Excel
    for index, (value_gave, value_get, value_course) in enumerate(zip(filtered_data_gara_history_swap['Отдал'],
                                                                      filtered_data_gara_history_swap['Получил'],
                                                                      filtered_data_gara_history_swap['Средняя цена'])):
        cell_gave = sheet.cell(row=cell_buy_gar_rub[0] + index, column=cell_buy_gar_rub[1])
        set_cell(cell_gave, value_gave)

        cell_get = sheet.cell(row=cell_buy_gar_usdt[0] + index, column=cell_buy_gar_usdt[1])
        set_cell(cell_get, value_get)

        cell_course = sheet.cell(row=cell_buy_gar_course[0] + index, column=cell_buy_gar_course[1])
        set_cell(cell_course, value_course)

        cell_market = sheet.cell(row=cell_buy_gar_market[0] + index, column=cell_buy_gar_market[1])
        set_cell(cell_market, 'Garantex')

    workbook.save(filename)


def write_to_excel_gara_account_history(filename, sheet, workbook, filtered_data_gara_history_account):
    cell_gar_comm = find_empty_cell_bottom(7, sheet)
    cell_gar_mean_price = find_empty_cell_bottom(6, sheet)
    cell_gar_multi = find_empty_cell_bottom(5, sheet)
    cell_gar_komsa_market = find_empty_cell_bottom(4, sheet)
    cell_gar_komsa = find_empty_cell_bottom(3, sheet)

    for index, (value_comm, value_mean_price) in enumerate(zip(filtered_data_gara_history_account['Комиссия'],
                                                               filtered_data_gara_history_account['Средняя цена'])):
        cell_comm = sheet.cell(row=cell_gar_comm[0] - index, column=cell_gar_comm[1])
        set_cell(cell_comm, (value_comm * -1))

        cell_mean_price = sheet.cell(row=cell_gar_mean_price[0] - index, column=cell_gar_mean_price[1])
        set_cell(cell_mean_price, value_mean_price)

        cell_multi = sheet.cell(row=cell_gar_multi[0] - index, column=cell_gar_multi[1])
        set_cell(cell_multi, ((value_comm * -1) * value_mean_price))

        cell_komsa_market = sheet.cell(row=cell_gar_komsa_market[0] - index, column=cell_gar_komsa_market[1])
        set_cell(cell_komsa_market, 'Garantex')

        cell_komsa = sheet.cell(row=cell_gar_komsa[0] - index, column=cell_gar_komsa[1])
        set_cell(cell_komsa, 'komsa')

    workbook.save(filename)


def write_to_excel_bybit_history_p2p(filename, sheet, workbook, filtered_data_bybit_history_p2p_buy_sum):
    cell_bybit_buy_usdt = find_empty_cell_top(7, sheet)
    cell_bybit_buy_mean = find_empty_cell_top(6, sheet)
    cell_bybit_buy_rub = find_empty_cell_top(5, sheet)
    cell_bybit_buy_market = find_empty_cell_top(4, sheet)

    for index, name_coin in enumerate(filtered_data_bybit_history_p2p_buy_sum.index):
        value_rub = filtered_data_bybit_history_p2p_buy_sum.loc[name_coin]['Fiat Amount']
        value_usdt = filtered_data_bybit_history_p2p_buy_sum.loc[name_coin]['Coin Amount']
        value_mean = filtered_data_bybit_history_p2p_buy_sum.loc[name_coin]['Price']
        if not name_coin == 'USDT':
            cell_bybit_buy_coin = (cell_bybit_buy_market[0], 3)
            cell_coin = sheet.cell(row=cell_bybit_buy_coin[0] + index, column=cell_bybit_buy_coin[1])
            set_cell(cell_coin, name_coin)

        cell_usdt = sheet.cell(row=cell_bybit_buy_usdt[0] + index, column=cell_bybit_buy_usdt[1])
        set_cell(cell_usdt, value_usdt)

        cell_rub = sheet.cell(row=cell_bybit_buy_rub[0] + index, column=cell_bybit_buy_rub[1])
        set_cell(cell_rub, value_rub)

        cell_mean = sheet.cell(row=cell_bybit_buy_mean[0] + index, column=cell_bybit_buy_mean[1])
        set_cell(cell_mean, value_mean)

        cell_market = sheet.cell(row=cell_bybit_buy_market[0] + index, column=cell_bybit_buy_market[1])
        set_cell(cell_market, 'Bybit')

    workbook.save(filename)


def write_to_excel_bybit_komsa(filename, sheet, workbook, filtered_data_bybit_for_komsa,
                               filtered_data_bybit_history_p2p_buy_sum):
    cell_bybit_comm = find_empty_cell_bottom(7, sheet)
    cell_bybit_mean_price = find_empty_cell_bottom(6, sheet)
    cell_bybit_mean_multi = find_empty_cell_bottom(5, sheet)
    cell_bybit_komsa_market = find_empty_cell_bottom(4, sheet)
    cell_bybit_komsa = find_empty_cell_bottom(3, sheet)

    cell_comm = sheet.cell(row=cell_bybit_comm[0], column=cell_bybit_comm[1])
    value_comm = filtered_data_bybit_for_komsa['Avg. Filled Price'].apply(
        lambda x: re.findall(r'\d+\.\d+', x)[0]).astype(float).sum() * -1
    set_cell(cell_comm, value_comm)

    cell_mean_price = sheet.cell(row=cell_bybit_mean_price[0], column=cell_bybit_mean_price[1])
    value_price_mean = filtered_data_bybit_history_p2p_buy_sum[filtered_data_bybit_history_p2p_buy_sum.index != 'USDT'][
        'Price'].sum()
    set_cell(cell_mean_price, value_price_mean)

    cell_mean_multi = sheet.cell(row=cell_bybit_mean_multi[0], column=cell_bybit_mean_multi[1])
    set_cell(cell_mean_multi, value_comm * value_price_mean)

    cell_komsa_market = sheet.cell(row=cell_bybit_komsa_market[0], column=cell_bybit_komsa_market[1])
    set_cell(cell_komsa_market, 'Bybit')

    cell_komsa = sheet.cell(row=cell_bybit_komsa[0], column=cell_bybit_komsa[1])
    set_cell(cell_komsa, 'komsa spot')

    workbook.save(filename)


def write_to_excel_exnode_history(filename, sheet, workbook, filtered_data_exnode_history):
    cell_exnode_buy_usdt = find_empty_cell_top(7, sheet)
    cell_exnode_buy_rub = find_empty_cell_top(5, sheet)
    cell_exnode_mean = find_empty_cell_top(6, sheet)
    cell_market = find_empty_cell_top(4, sheet)

    for index, (value_usdt, value_rub) in enumerate(
            zip(filtered_data_exnode_history.loc[filtered_data_exnode_history['direction'] == 'IN', 'usdt'],
                filtered_data_exnode_history.loc[filtered_data_exnode_history['direction'] == 'IN', 'rub'])):
        cell_usdt = sheet.cell(row=cell_exnode_buy_usdt[0] + index, column=cell_exnode_buy_usdt[1])
        set_cell(cell_usdt, value_usdt)

        cell_rub = sheet.cell(row=cell_exnode_buy_rub[0] + index, column=cell_exnode_buy_rub[1])
        set_cell(cell_rub, value_rub)

        cell_rub = sheet.cell(row=cell_exnode_mean[0] + index, column=cell_exnode_mean[1])
        set_cell(cell_rub, value_rub / value_usdt)

        cell_rub = sheet.cell(row=cell_market[0] + index, column=cell_market[1])
        set_cell(cell_rub, 'EXNODE')

    workbook.save(filename)


def write_to_excel_exnode_komsa(filename, sheet, workbook, filtered_data_exnode_history):
    cell_exnode_buy_comm = find_empty_cell_bottom(7, sheet)
    cell_exnode_mean_price = find_empty_cell_bottom(6, sheet)
    cell_exnode_multi = find_empty_cell_bottom(5, sheet)
    cell_exnode_komsa_market = find_empty_cell_bottom(4, sheet)
    cell_exnode_komsa = find_empty_cell_bottom(3, sheet)

    value_mean_price = round(filtered_data_exnode_history[filtered_data_exnode_history['direction'] == 'IN'].rub.sum() /
                             filtered_data_exnode_history[filtered_data_exnode_history['direction'] == 'IN'].usdt.sum(),
                             2)

    for index, (value_usdt, value_comm) in enumerate(
            zip(filtered_data_exnode_history.loc[filtered_data_exnode_history['direction'] == 'OUT', 'usdt'],
                filtered_data_exnode_history.loc[filtered_data_exnode_history['direction'] == 'OUT', 'commission'])):
        cell_comm = sheet.cell(row=cell_exnode_buy_comm[0] + index, column=cell_exnode_buy_comm[1])
        set_cell(cell_comm, value_comm * -1)

        cell_exnode_mean_price = sheet.cell(row=cell_exnode_mean_price[0] + index, column=cell_exnode_mean_price[1])
        set_cell(cell_exnode_mean_price, value_mean_price)

        cell_multi = sheet.cell(row=cell_exnode_multi[0] + index, column=cell_exnode_multi[1])
        set_cell(cell_multi, value_mean_price * (value_comm * -1))

        cell_komsa_market = sheet.cell(row=cell_exnode_komsa_market[0] + index, column=cell_exnode_komsa_market[1])
        set_cell(cell_komsa_market, 'EXNODE')

        cell_komsa = sheet.cell(row=cell_exnode_komsa[0] + index, column=cell_exnode_komsa[1])
        set_cell(cell_komsa, 'komsa')

    workbook.save(filename)


def write_to_to_excel_huobi_history_p2p_buy(filename, sheet, workbook, huobi_history_p2p_buy_sum_by_coin,
                                            huobi_spot_mean_sum):
    cell_huobi_buy_usdt = find_empty_cell_top(7, sheet)
    cell_huobi_buy_mean = find_empty_cell_top(6, sheet)
    cell_huobi_buy_rub = find_empty_cell_top(5, sheet)
    cell_huobi_buy_market = find_empty_cell_top(4, sheet)

    for index, name_coin in enumerate(huobi_history_p2p_buy_sum_by_coin.index):
        value_rub = huobi_history_p2p_buy_sum_by_coin.loc[name_coin]['Общая цена']
        if not name_coin == 'USDT':
            value_usdt = (huobi_spot_mean_sum[huobi_spot_mean_sum['Пара'].str.contains(name_coin)]['Цена'] *
                          huobi_history_p2p_buy_sum_by_coin.loc[name_coin]['Количество']).values[0]
            huobi_spot_mean_sum.loc[
                huobi_spot_mean_sum['Пара'].str.contains(name_coin), 'usdt/rub'] = value_rub / value_usdt
            cell_coin = (cell_huobi_buy_market[0], 3)
            cell_market = sheet.cell(row=cell_coin[0] + index, column=cell_coin[1])
            set_cell(cell_market, name_coin)
        else:
            value_usdt = huobi_history_p2p_buy_sum_by_coin.loc[name_coin]['Количество']

        value_mean = value_rub / value_usdt
        huobi_history_p2p_buy_sum_by_coin.loc[name_coin, 'usdt/rub'] = value_mean

        cell_usdt = sheet.cell(row=cell_huobi_buy_usdt[0] + index, column=cell_huobi_buy_usdt[1])
        set_cell(cell_usdt, value_usdt)

        cell_rub = sheet.cell(row=cell_huobi_buy_rub[0] + index, column=cell_huobi_buy_rub[1])
        set_cell(cell_rub, value_rub)

        cell_mean = sheet.cell(row=cell_huobi_buy_mean[0] + index, column=cell_huobi_buy_mean[1])
        set_cell(cell_mean, value_mean)

        cell_market = sheet.cell(row=cell_huobi_buy_market[0] + index, column=cell_huobi_buy_market[1])
        set_cell(cell_market, 'huobi')

        value_rub, value_usdt = 0, 0

    workbook.save(filename)


def write_to_excel_huobi_komsa(filename, sheet, workbook, huobi_spot_mean_sum):
    cell_huobi_komsa = find_empty_cell_bottom(7, sheet)
    cell_huobi_komsa_mean = find_empty_cell_bottom(6, sheet)
    cell_huobi_komsa_rub = find_empty_cell_bottom(5, sheet)
    cell_huobi_komsa_market = find_empty_cell_bottom(4, sheet)
    cell_huobi_komsa_name = find_empty_cell_bottom(3, sheet)

    huobi_spot_mean_sum.set_index('Пара', inplace=True)
    for index, value in enumerate(huobi_spot_mean_sum.index):
        cell_komsa = sheet.cell(row=cell_huobi_komsa[0] + index, column=cell_huobi_komsa[1])
        value_komsa = huobi_spot_mean_sum.loc[value]['Комиссия'] * -1
        set_cell(cell_komsa, value_komsa)

        cell_komsa_mean = sheet.cell(row=cell_huobi_komsa_mean[0] + index, column=cell_huobi_komsa_mean[1])
        value_komsa_mean = huobi_spot_mean_sum.loc[value]['usdt/rub']
        set_cell(cell_komsa_mean, value_komsa_mean)

        cell_komsa_rub = sheet.cell(row=cell_huobi_komsa_rub[0] + index, column=cell_huobi_komsa_rub[1])
        set_cell(cell_komsa_rub, value_komsa_mean * value_komsa)

        cell_market = sheet.cell(row=cell_huobi_komsa_market[0] + index, column=cell_huobi_komsa_market[1])
        set_cell(cell_market, 'huobi')

        cell_komsa_name = sheet.cell(row=cell_huobi_komsa_name[0] + index, column=cell_huobi_komsa_name[1])
        set_cell(cell_komsa_name, 'komsa spot')

    workbook.save(filename)
